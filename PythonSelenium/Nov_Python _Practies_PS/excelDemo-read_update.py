import openpyxl
Dict ={}
book = openpyxl.load_workbook("C:\\pythonexcel\\python_excel.xlsx")
sheet = book.active
#cell = sheet.cell(row=1, column=2)
#print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"
print(sheet.cell(row=2, column=2).value)


for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase4":
        for j in range(1, sheet.max_column + 1):
            key = sheet.cell(row=1, column=j).value
            value = sheet.cell(row=i, column=j).value
            Dict[key] = value

print(Dict)


