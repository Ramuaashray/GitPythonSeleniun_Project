import openpyxl
from selenium.webdriver.common.by import By

from PageObjects.file_Validation import file_validation
from Old_Project_Created_Files.Upload_File import update_excel_data, fruit_name, newValue


class BrowserUtils:
    pass


class file_download_update(BrowserUtils):

    def __init__(self, driver):
        self.driver = driver
        self.download_button = (By.ID, "downloadButton")
        self.file_input = (By.CSS_SELECTOR, "input[type='file']")

        global file_path
        file_path = "C:\\Users\\MERAMU\\Downloads\\download.xlsx"




    def update_excel_data(filePath, searchTerm, colName, new_value):
        book = openpyxl.load_workbook(filePath)
        sheet = book.active
        Dict = {}

        for i in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=i).value == colName:
                Dict["col"] = i

        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                if sheet.cell(row=i, column=j).value == searchTerm:
                    Dict["row"] = i

        sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
        book.save(file_path)

    def download_file(self):
        self.driver.find_element(*self.download_button ).click()

    def update_file(self):
        # edit the excel with updated value
        update_excel_data(file_path, fruit_name, "price", newValue)

    def upload_file(self):
        # upload
        file_input = self.driver.find_element(*self.file_input)
        file_input.send_keys(file_path)
        file_vali = file_validation
        return file_vali



